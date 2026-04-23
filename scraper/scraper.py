"""
QUBIE NEWS Scraper
===================
Fetches articles from RSS feeds, search APIs, and social/forum platforms,
scores them by quantum biology keyword relevance, and writes:
    - ../feed.json                       (for the frontend)
    - ../QBIO-Report-Sources.xlsx        (source inventory with live stats)

Usage:
    python scraper.py
"""

import json
import os
import re
import time
import urllib.parse
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime

import feedparser
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Optional: .env + Anthropic SDK for LLM blurbs. Gracefully degrade if unavailable.
try:
    from dotenv import load_dotenv
    # override=True so our .env always wins over any empty / stale system env var
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"), override=True)
except ImportError:
    pass

try:
    import anthropic
except ImportError:
    anthropic = None


# ============================================================================
# CONFIGURATION
# ============================================================================

# --- RSS Feeds (Tier 1) --------------------------------------------------
RSS_FEEDS = [
    ("arXiv — Quantum Physics",       "https://rss.arxiv.org/rss/quant-ph",                                  "preprint"),
    ("arXiv — Quantitative Biology",  "https://rss.arxiv.org/rss/q-bio",                                     "preprint"),
    ("bioRxiv — Biophysics",          "http://connect.biorxiv.org/biorxiv_xml.php?subject=biophysics",       "preprint"),
    ("bioRxiv — Biochemistry",        "http://connect.biorxiv.org/biorxiv_xml.php?subject=biochemistry",     "preprint"),
    ("bioRxiv — Molecular Biology",   "http://connect.biorxiv.org/biorxiv_xml.php?subject=molecular_biology","preprint"),
    ("bioRxiv — Systems Biology",     "http://connect.biorxiv.org/biorxiv_xml.php?subject=systems_biology",  "preprint"),
    ("bioRxiv — Neuroscience",        "http://connect.biorxiv.org/biorxiv_xml.php?subject=neuroscience",     "preprint"),
    # Dropped: ChemRxiv and Royal Society Interface (both Cloudflare 403).
    # Their papers are still indexed via Europe PMC (ChemRxiv) and PubMed (RSIF),
    # both of which we already query — so coverage isn't lost.
    ("Nature Physics",                "https://www.nature.com/nphys.rss",                                    "paper"),
    ("Nature Chemistry",              "https://www.nature.com/nchem.rss",                                    "paper"),
    ("Nature",                        "https://www.nature.com/nature.rss",                                   "paper"),
    ("Science Magazine",              "https://www.science.org/action/showFeed?type=etoc&feed=rss&jc=science","paper"),
    ("PNAS",                          "https://www.pnas.org/action/showFeed?type=etoc&feed=rss&jc=pnas",     "paper"),
    ("Quanta Magazine",               "https://www.quantamagazine.org/feed/",                                "news"),
    ("New Scientist — Physics",       "https://www.newscientist.com/subject/physics/feed/",                  "news"),
    ("Science News",                  "https://www.sciencenews.org/feed",                                    "news"),
    ("Phys.org — Physics",            "https://phys.org/rss-feed/physics-news/",                             "news"),
    ("ScienceDaily",                  "https://www.sciencedaily.com/rss/all.xml",                            "news"),
]

# --- Reddit subreddits (Tier 3a) ----------------------------------------
# Reddit doesn't use keyword queries — we pull recent posts and score them.
REDDIT_SUBREDDITS = [
    "quantumbiology",
    "biophysics",
    "askscience",
    "Physics",
    "biochemistry",
]

# --- Stack Exchange (Tier 3a) -------------------------------------------
# KEPT CURATED (not from keywords.txt) because SE has a 300 req/day quota
# without an API key. Using all 46 keywords * 3 sites would burn 46% of quota
# per scraper run. Each tuple: (site_id, search_query)
STACK_EXCHANGE = [
    ("biology",   "quantum"),
    ("physics",   "biology"),
    ("chemistry", "quantum biology"),
]

# --- YouTube channels (Tier 4) -----------------------------------------
# Each entry: (channel_display_name, channel_id). The channel ID starts with
# "UC" and can be found at youtube.com/channel/UC... or by visiting the
# channel's About page on YT. Channel RSS is unauth'd and unlimited.
YOUTUBE_CHANNELS = [
    # Initial 3 channels (Royal Institution / Sabine / PBS Space Time) were
    # broad physics/science creators — the keyword filter dropped 100% of
    # their videos. Leaving this list empty for now; Tier 4 relies on the
    # keyword-search API alone. Add quantum-biology-specific creators here
    # as we find them (tuple: (display_name, channel_id starting with "UC")).
]

# --- Merge in runtime sources config (pushed via /admin's "Push" button) ---
# Admins can push new RSS feeds or subreddits through the UI; those get
# appended to sources_config.json on the volume. We merge them here so the
# scraper picks them up on the very next run.
def _merge_runtime_sources():
    global RSS_FEEDS, REDDIT_SUBREDDITS
    path = SOURCES_CONFIG_FILE
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except (json.JSONDecodeError, OSError) as e:
        print(f"[runtime-sources] couldn't parse {path}: {e}")
        return

    added_feeds = 0
    for item in cfg.get("rss_feeds", []):
        name = (item.get("name") or "").strip()
        url  = (item.get("url") or "").strip()
        cat  = (item.get("category") or "news").strip()
        if name and url:
            RSS_FEEDS.append((name, url, cat))
            added_feeds += 1

    added_subs = 0
    existing_subs_lower = {s.lower() for s in REDDIT_SUBREDDITS}
    for sub in cfg.get("reddit_subreddits", []):
        s = (sub or "").strip().lstrip("r/").lstrip("/r/")
        if s and s.lower() not in existing_subs_lower:
            REDDIT_SUBREDDITS.append(s)
            existing_subs_lower.add(s.lower())
            added_subs += 1

    if added_feeds or added_subs:
        print(f"[runtime-sources] merged from {path}: "
              f"+{added_feeds} RSS feed(s), +{added_subs} subreddit(s)")


# How many keywords to OR together into a single API call.
# (Avoids URL-length limits; PubMed/arXiv/EuropePMC all handle chunked queries fine.)
KEYWORD_CHUNK_SIZE = 10

# --- File paths ----------------------------------------------------------
# Locally:    files live alongside scraper.py (../)
# On Railway: DATA_DIR env var points to the persistent volume (/data)
HERE          = os.path.dirname(__file__)
DATA_DIR      = os.environ.get("DATA_DIR") or os.path.abspath(os.path.join(HERE, ".."))
OUTPUT_FILE   = os.path.join(DATA_DIR, "feed.json")
XLSX_FILE     = os.path.join(DATA_DIR, "QBIO-Report-Sources.xlsx")
SOURCES_JSON  = os.path.join(DATA_DIR, "sources.json")
KEYWORDS_FILE = os.path.join(DATA_DIR, "keywords.txt")
# User-pushed sources (runtime overrides). The server's Admin page writes here
# when someone clicks "Push" on a source request. Merged with hardcoded
# RSS_FEEDS / REDDIT_SUBREDDITS at scrape time.
SOURCES_CONFIG_FILE = os.path.join(DATA_DIR, "sources_config.json")
# Fallback to bundled keywords.txt if the data dir doesn't have one yet
# (first run on a fresh volume).
if not os.path.exists(KEYWORDS_FILE):
    bundled = os.path.join(HERE, "keywords.txt")
    if os.path.exists(bundled):
        KEYWORDS_FILE = bundled

# Now that SOURCES_CONFIG_FILE is defined, actually merge any runtime overrides.
_merge_runtime_sources()

# --- Request settings ----------------------------------------------------
REQUEST_TIMEOUT = 20
USER_AGENT      = "QBIO-Report/1.0 (contact: ollipayne182@gmail.com)"
HEADERS         = {"User-Agent": USER_AGENT}
API_DELAY       = 0.4

# Reddit requires a distinctive, app-identifying User-Agent. They block
# generic UAs and cloud-provider IPs unless requests are OAuth-authenticated.
REDDIT_UA       = "QUBIE-News-scraper/1.0 (contact: ollipayne182@gmail.com)"

MAX_PER_CALL = 50


# ============================================================================
# STATS TRACKING (for the auto-generated xlsx)
# ============================================================================
# Every fetcher appends one row here describing its result.
# Each row: {tier, source_name, target, category, matched, total, status, error, notes}

STATS_ROWS = []

def record_stats(tier, source_name, target, category, matched, total, error="", notes=""):
    """Record a per-source result for the xlsx output."""
    if error:
        status = "ERROR"
    elif total == 0:
        status = "No results returned"
    elif matched == 0:
        status = "Working (no hits)"
    else:
        status = "Working"
    STATS_ROWS.append({
        "tier": tier, "source_name": source_name, "target": target,
        "category": category, "matched": matched, "total": total,
        "status": status, "error": error, "notes": notes,
    })


# ============================================================================
# SHARED HELPERS
# ============================================================================

_WEIGHT_RE = re.compile(r"^(.*?)\s*\[(\d+)\]\s*$")

def load_keywords(filepath=KEYWORDS_FILE):
    """Parse keywords.txt. Returns dict {phrase: weight}.
    Default weight is 1. Append `[N]` to a line to assign weight N.
    Example: 'quantum biology [5]' -> weight 5.
    """
    weights = {}
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            m = _WEIGHT_RE.match(line)
            if m:
                phrase = m.group(1).strip().lower()
                weight = int(m.group(2))
            else:
                phrase = line.lower()
                weight = 1
            if phrase:
                weights[phrase] = weight
    weighted = sum(1 for w in weights.values() if w > 1)
    print(f"Loaded {len(weights)} keywords from {filepath} ({weighted} weighted > 1)")
    return weights


def clean_html(raw_html):
    return re.sub(r"<[^>]+>", "", raw_html or "").strip()


def score_article(title, summary, keywords):
    """Score = sum of matched keyword weights. Returns (score, matched_list).
    `keywords` is a dict {phrase: weight}; a plain list/iterable is also
    accepted for backwards compatibility (treats each as weight 1)."""
    text = f"{title} {summary}".lower()
    if isinstance(keywords, dict):
        matched = []
        score = 0
        for phrase, weight in keywords.items():
            if phrase in text:
                matched.append(phrase)
                score += weight
        return score, matched
    matched = [kw for kw in keywords if kw in text]
    return len(matched), matched


def chunk_list(items, size):
    """Chunk a list, iterable, or dict (yields keys in chunks for dicts)."""
    items = list(items) if not isinstance(items, list) else items
    for i in range(0, len(items), size):
        yield items[i:i + size]


def parse_date_to_iso(date_str):
    if not date_str:
        return ""
    date_str = date_str.strip()
    try:
        dt = parsedate_to_datetime(date_str)
        if dt:
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc).isoformat()
    except (TypeError, ValueError):
        pass
    for candidate in [date_str, date_str.replace("Z", "+00:00")]:
        try:
            dt = datetime.fromisoformat(candidate)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc).isoformat()
        except ValueError:
            pass
    for fmt in ["%Y %b %d", "%Y %b", "%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%d %b %Y", "%Y"]:
        try:
            dt = datetime.strptime(date_str, fmt).replace(tzinfo=timezone.utc)
            return dt.isoformat()
        except ValueError:
            pass
    return ""


def make_article(title, link, source, category, date, score, matched, summary, thumbnail=""):
    """`thumbnail` is an optional URL (populated for videos)."""
    return {
        "title":             title,
        "link":              link,
        "source":            source,
        "source_category":   category,
        "date":              date,
        "date_iso":          parse_date_to_iso(date),
        "score":             score,
        "matched_keywords":  matched,
        "summary":           summary[:500] if summary else "",
        "thumbnail":         thumbnail or "",
    }


def build_or_query(terms, quote=True):
    if quote:
        return " OR ".join(f'"{t}"' for t in terms)
    return " OR ".join(terms)


# ============================================================================
# FETCHER: RSS Feeds
# ============================================================================

def fetch_rss(name, url, category, keywords):
    print(f"  RSS: {name} ...")
    error = ""
    entries_count = 0
    articles = []
    try:
        response = requests.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        # Detect bot-blocked responses (Cloudflare etc.) — they return HTTP
        # 403/503 with an HTML challenge page instead of a real feed.
        if response.status_code >= 400:
            raise Exception(f"HTTP {response.status_code}")
        body_head = response.content[:200].lower()
        if b"<!doctype html" in body_head or b"<html" in body_head:
            raise Exception("Bot-blocked (Cloudflare or similar) - HTML returned instead of feed")
        feed = feedparser.parse(response.content)
        entries_count = len(feed.entries)
        for entry in feed.entries[:MAX_PER_CALL]:
            title   = clean_html(getattr(entry, "title", ""))
            summary = clean_html(getattr(entry, "summary", ""))
            link    = getattr(entry, "link", "")
            date    = getattr(entry, "published", "") or getattr(entry, "updated", "")
            score, matched = score_article(title, summary, keywords)
            if score > 0:
                articles.append(make_article(title, link, name, category, date, score, matched, summary))
    except Exception as e:
        error = str(e)
        print(f"    ERROR: {error}")

    record_stats("Tier 1", name, url, category, len(articles), entries_count, error)
    print(f"    -> {len(articles)} matched (of {entries_count})")
    return articles


# ============================================================================
# FETCHER: PubMed (chunked OR queries, uses ALL keywords)
# ============================================================================

def fetch_pubmed(keywords):
    print("  API: PubMed ...")
    all_articles = []
    error = ""
    total_seen = 0

    for chunk in chunk_list(keywords, KEYWORD_CHUNK_SIZE):
        query = build_or_query(chunk)
        try:
            r = requests.get(
                "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi",
                params={"db": "pubmed", "term": query, "retmode": "json",
                        "retmax": MAX_PER_CALL, "sort": "pub_date"},
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            ids = r.json().get("esearchresult", {}).get("idlist", [])
        except Exception as e:
            error = str(e)
            print(f"    ERROR (esearch): {error}")
            break
        if not ids:
            continue
        time.sleep(API_DELAY)
        total_seen += len(ids)
        try:
            r = requests.get(
                "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi",
                params={"db": "pubmed", "id": ",".join(ids), "retmode": "json"},
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            result = r.json().get("result", {})
        except Exception as e:
            error = str(e)
            print(f"    ERROR (esummary): {error}")
            continue
        for pmid in ids:
            data = result.get(pmid)
            if not data:
                continue
            title   = data.get("title", "")
            journal = data.get("fulljournalname") or data.get("source", "")
            date    = data.get("pubdate", "")
            link    = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            score, matched = score_article(title, journal, keywords)
            if score > 0:
                all_articles.append(make_article(title, link, f"PubMed · {journal}", "paper", date, score, matched, ""))
        time.sleep(API_DELAY)

    record_stats("Tier 2", "PubMed E-utilities",
                 "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/",
                 "paper", len(all_articles), total_seen, error,
                 notes=f"Chunked into {len(list(chunk_list(keywords, KEYWORD_CHUNK_SIZE)))} API calls")
    print(f"    -> {len(all_articles)} matched (of {total_seen} PMIDs across chunks)")
    return all_articles


# ============================================================================
# FETCHER: arXiv Search API (chunked OR queries, uses ALL keywords)
# ============================================================================

def fetch_arxiv_api(keywords):
    print("  API: arXiv search ...")
    all_articles = []
    error = ""
    total_seen = 0
    seen_links = set()

    for chunk in chunk_list(keywords, KEYWORD_CHUNK_SIZE):
        query_terms = " OR ".join(f'all:"{t}"' for t in chunk)
        try:
            r = requests.get(
                "http://export.arxiv.org/api/query",
                params={
                    "search_query": query_terms,
                    "max_results": MAX_PER_CALL,
                    "sortBy": "submittedDate",
                    "sortOrder": "descending",
                },
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            feed = feedparser.parse(r.content)
        except Exception as e:
            error = str(e)
            print(f"    ERROR: {error}")
            break
        total_seen += len(feed.entries)
        for entry in feed.entries:
            link = getattr(entry, "link", "")
            if link in seen_links:
                continue
            seen_links.add(link)
            title   = clean_html(getattr(entry, "title", ""))
            summary = clean_html(getattr(entry, "summary", ""))
            date    = getattr(entry, "published", "")
            score, matched = score_article(title, summary, keywords)
            if score > 0:
                all_articles.append(make_article(title, link, "arXiv (search)", "preprint", date, score, matched, summary))
        time.sleep(API_DELAY)

    record_stats("Tier 2", "arXiv Search API",
                 "http://export.arxiv.org/api/query",
                 "preprint", len(all_articles), total_seen, error,
                 notes=f"Chunked into {len(list(chunk_list(keywords, KEYWORD_CHUNK_SIZE)))} API calls")
    print(f"    -> {len(all_articles)} matched (of {total_seen} results across chunks)")
    return all_articles


# ============================================================================
# FETCHER: Europe PMC (chunked OR queries, uses ALL keywords)
# ============================================================================

def fetch_europepmc(keywords):
    print("  API: Europe PMC ...")
    all_articles = []
    error = ""
    total_seen = 0
    seen_links = set()

    for chunk in chunk_list(keywords, KEYWORD_CHUNK_SIZE):
        query = build_or_query(chunk)
        try:
            r = requests.get(
                "https://www.ebi.ac.uk/europepmc/webservices/rest/search",
                params={"query": query, "format": "json",
                        "resultType": "lite", "pageSize": MAX_PER_CALL,
                        # Europe PMC's sort requires their own field syntax.
                        # Plain "date" silently returns zero results.
                        "sort": "FIRST_PDATE_D desc"},
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            results = r.json().get("resultList", {}).get("result", [])
        except Exception as e:
            error = str(e)
            print(f"    ERROR: {error}")
            break
        total_seen += len(results)
        for item in results:
            title   = item.get("title", "")
            journal = item.get("journalTitle", "")
            doi     = item.get("doi", "")
            pmid    = item.get("pmid", "")
            pmcid   = item.get("pmcid", "")
            date    = item.get("firstPublicationDate", "") or item.get("pubYear", "")
            if doi:     link = f"https://doi.org/{doi}"
            elif pmid:  link = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            elif pmcid: link = f"https://europepmc.org/article/PMC/{pmcid}"
            else:       link = ""
            if not link or link in seen_links:
                continue
            seen_links.add(link)
            score, matched = score_article(title, journal, keywords)
            if score > 0:
                source_name = f"Europe PMC · {journal}" if journal else "Europe PMC"
                all_articles.append(make_article(title, link, source_name, "paper", date, score, matched, ""))
        time.sleep(API_DELAY)

    record_stats("Tier 2", "Europe PMC",
                 "https://www.ebi.ac.uk/europepmc/webservices/rest/search",
                 "paper", len(all_articles), total_seen, error,
                 notes=f"Chunked into {len(list(chunk_list(keywords, KEYWORD_CHUNK_SIZE)))} API calls")
    print(f"    -> {len(all_articles)} matched (of {total_seen} results across chunks)")
    return all_articles


# ============================================================================
# FETCHER: Semantic Scholar (~200M papers, free API, no key required)
# ============================================================================

def fetch_semantic_scholar(keywords):
    """Search Semantic Scholar for papers matching our keywords.
    Without an API key, their public tier is very aggressively rate-limited
    from cloud IPs (~0 reliable results per scrape). Gated on env var
    SEMANTIC_SCHOLAR_API_KEY; skip entirely if not set.
    Apply for a free key: https://www.semanticscholar.org/product/api#api-key-form
    """
    api_key = os.environ.get("SEMANTIC_SCHOLAR_API_KEY")
    if not api_key:
        print("  API: Semantic Scholar ... skipped (no SEMANTIC_SCHOLAR_API_KEY set)")
        record_stats("Tier 2", "Semantic Scholar",
                     "https://api.semanticscholar.org/graph/v1/paper/search",
                     "paper", 0, 0, "",
                     notes="Skipped — apply for free API key at semanticscholar.org/product/api")
        return []

    print("  API: Semantic Scholar ...")
    all_articles = []
    error = ""
    total_seen = 0
    seen_links = set()
    headers_with_key = {**HEADERS, "x-api-key": api_key}

    for chunk in chunk_list(keywords, KEYWORD_CHUNK_SIZE):
        # Semantic Scholar's `query` is natural-language relevance search; join
        # chunk phrases with spaces and let their ranker sort it out.
        query = " ".join(chunk)
        try:
            r = requests.get(
                "https://api.semanticscholar.org/graph/v1/paper/search",
                params={
                    "query": query,
                    "limit": MAX_PER_CALL,
                    "fields": "title,abstract,url,externalIds,publicationDate,venue,year",
                },
                timeout=REQUEST_TIMEOUT, headers=headers_with_key,
            )
            if r.status_code == 429:
                print(f"    rate-limited, backing off")
                time.sleep(5)
                continue
            if r.status_code != 200:
                raise Exception(f"HTTP {r.status_code}")
            results = r.json().get("data", []) or []
        except Exception as e:
            error = str(e)
            print(f"    ERROR: {error}")
            break
        total_seen += len(results)
        for item in results:
            title    = item.get("title") or ""
            abstract = item.get("abstract") or ""
            ext_ids  = item.get("externalIds") or {}
            # Prefer DOI > arXiv > S2 URL
            link = ""
            if ext_ids.get("DOI"):
                link = f"https://doi.org/{ext_ids['DOI']}"
            elif ext_ids.get("ArXiv"):
                link = f"https://arxiv.org/abs/{ext_ids['ArXiv']}"
            elif item.get("url"):
                link = item["url"]
            if not link or link in seen_links:
                continue
            seen_links.add(link)
            date  = item.get("publicationDate") or (str(item["year"]) + "-01-01" if item.get("year") else "")
            venue = item.get("venue") or ""
            score, matched = score_article(title, abstract, keywords)
            if score > 0:
                source_name = f"Semantic Scholar · {venue}" if venue else "Semantic Scholar"
                all_articles.append(make_article(title, link, source_name, "paper", date, score, matched, abstract))
        time.sleep(API_DELAY * 3)  # Stricter rate limit without API key

    record_stats("Tier 2", "Semantic Scholar",
                 "https://api.semanticscholar.org/graph/v1/paper/search",
                 "paper", len(all_articles), total_seen, error,
                 notes=f"Chunked into {len(list(chunk_list(keywords, KEYWORD_CHUNK_SIZE)))} API calls")
    print(f"    -> {len(all_articles)} matched (of {total_seen} results across chunks)")
    return all_articles


# ============================================================================
# FETCHER: OpenAlex (250M+ scholarly works, free API, no key)
# ============================================================================

def _reconstruct_abstract(inverted_index):
    """OpenAlex returns abstracts as {word: [positions]} — rebuild the string."""
    if not inverted_index:
        return ""
    try:
        words_at = {}
        for word, positions in inverted_index.items():
            for p in positions:
                words_at[p] = word
        return " ".join(words_at[i] for i in sorted(words_at))
    except Exception:
        return ""


def fetch_openalex(keywords):
    """Search OpenAlex for works matching our keywords. Free, no key needed."""
    print("  API: OpenAlex ...")
    all_articles = []
    error = ""
    total_seen = 0
    seen_links = set()

    for chunk in chunk_list(keywords, KEYWORD_CHUNK_SIZE):
        query = " ".join(chunk)
        try:
            r = requests.get(
                "https://api.openalex.org/works",
                params={
                    "search": query,
                    "per-page": MAX_PER_CALL,
                    "sort": "publication_date:desc",
                    # Email in `mailto` puts us in OpenAlex's "polite pool" (faster)
                    "mailto": "ollipayne182@gmail.com",
                },
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            if r.status_code != 200:
                raise Exception(f"HTTP {r.status_code}")
            results = r.json().get("results", []) or []
        except Exception as e:
            error = str(e)
            print(f"    ERROR: {error}")
            break
        total_seen += len(results)
        for item in results:
            title    = item.get("title") or item.get("display_name") or ""
            abstract = _reconstruct_abstract(item.get("abstract_inverted_index"))
            doi      = item.get("doi") or ""
            # doi is already a URL like "https://doi.org/10.xxxx"
            link     = doi or item.get("id") or ""
            if not link or link in seen_links:
                continue
            seen_links.add(link)
            date = item.get("publication_date") or (
                str(item["publication_year"]) + "-01-01" if item.get("publication_year") else ""
            )
            primary_loc = item.get("primary_location") or {}
            venue = (primary_loc.get("source") or {}).get("display_name", "") if primary_loc else ""
            score, matched = score_article(title, abstract, keywords)
            if score > 0:
                source_name = f"OpenAlex · {venue}" if venue else "OpenAlex"
                all_articles.append(make_article(title, link, source_name, "paper", date, score, matched, abstract))
        time.sleep(API_DELAY)

    record_stats("Tier 2", "OpenAlex",
                 "https://api.openalex.org/works",
                 "paper", len(all_articles), total_seen, error,
                 notes=f"Chunked into {len(list(chunk_list(keywords, KEYWORD_CHUNK_SIZE)))} API calls")
    print(f"    -> {len(all_articles)} matched (of {total_seen} results across chunks)")
    return all_articles


# ============================================================================
# FETCHER: YouTube Data API — keyword search across all of YouTube
# ============================================================================
# Requires YOUTUBE_API_KEY env var. Free quota: 10,000 units/day; search.list
# costs 100 units per call. With our chunked keyword approach that's ~5 calls
# per scrape = 500 units = well under quota.

def fetch_youtube_api(keywords):
    """Keyword-search videos on YouTube via the Data API v3."""
    api_key = os.environ.get("YOUTUBE_API_KEY")
    if not api_key:
        print("  API: YouTube search ... skipped (no YOUTUBE_API_KEY set)")
        record_stats("Tier 4", "YouTube Data API", "search.list",
                     "video", 0, 0, "",
                     notes="Skipped — get a free key at console.cloud.google.com")
        return []

    print("  API: YouTube search ...")
    all_articles = []
    error = ""
    total_seen = 0
    seen_links = set()

    for chunk in chunk_list(keywords, KEYWORD_CHUNK_SIZE):
        # YouTube's `q` supports `|` as OR and `""` for phrase matching.
        # Space-joining phrases would make YT require ALL words simultaneously
        # (0 results). OR-joining quoted phrases surfaces videos matching any.
        query = " | ".join(f'"{kw}"' if " " in kw else kw for kw in chunk)
        try:
            r = requests.get(
                "https://www.googleapis.com/youtube/v3/search",
                params={
                    "key": api_key,
                    "q": query,
                    "part": "snippet",
                    "type": "video",
                    "maxResults": 50,        # YT API max per page
                    "order": "date",
                    "relevanceLanguage": "en",
                    "safeSearch": "none",
                },
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            if r.status_code != 200:
                raise Exception(f"HTTP {r.status_code}: {r.text[:160]}")
            items = r.json().get("items", []) or []
        except Exception as e:
            error = str(e)
            print(f"    ERROR: {error}")
            break
        total_seen += len(items)
        for item in items:
            video_id = (item.get("id") or {}).get("videoId", "")
            snip     = item.get("snippet") or {}
            if not video_id: continue
            link = f"https://www.youtube.com/watch?v={video_id}"
            if link in seen_links: continue
            seen_links.add(link)

            title       = snip.get("title", "")
            desc        = snip.get("description", "")
            channel     = snip.get("channelTitle", "")
            published   = snip.get("publishedAt", "")
            thumbs      = snip.get("thumbnails") or {}
            thumb_url   = ((thumbs.get("high") or thumbs.get("medium") or thumbs.get("default")) or {}).get("url", "")
            score, matched = score_article(title, desc, keywords)
            if score > 0:
                source_name = f"YouTube · {channel}" if channel else "YouTube"
                all_articles.append(make_article(
                    title, link, source_name, "video", published, score, matched, desc,
                    thumbnail=thumb_url,
                ))
        time.sleep(API_DELAY)

    record_stats("Tier 4", "YouTube Data API",
                 "https://www.googleapis.com/youtube/v3/search",
                 "video", len(all_articles), total_seen, error,
                 notes=f"Chunked into {len(list(chunk_list(keywords, KEYWORD_CHUNK_SIZE)))} API calls (~100 quota units each)")
    print(f"    -> {len(all_articles)} matched (of {total_seen} videos across chunks)")
    return all_articles


# ============================================================================
# FETCHER: YouTube channel RSS — follow specific creators (free, no auth)
# ============================================================================

def fetch_youtube_channel(channel_name, channel_id, keywords):
    """Fetch a specific YouTube channel's RSS feed."""
    print(f"  YouTube RSS: {channel_name} ...")
    url = f"https://www.youtube.com/feeds/videos.xml?channel_id={channel_id}"
    error = ""
    entries_count = 0
    articles = []
    try:
        response = requests.get(url, timeout=REQUEST_TIMEOUT, headers=HEADERS)
        if response.status_code >= 400:
            raise Exception(f"HTTP {response.status_code}")
        feed = feedparser.parse(response.content)
        entries_count = len(feed.entries)
        for entry in feed.entries[:MAX_PER_CALL]:
            title = clean_html(getattr(entry, "title", ""))
            link  = getattr(entry, "link", "")
            date  = getattr(entry, "published", "") or getattr(entry, "updated", "")
            # YouTube RSS: description is under media_description or summary
            desc  = clean_html(getattr(entry, "summary", "") or "")
            # Thumbnail: media:thumbnail
            thumb = ""
            mthumbs = getattr(entry, "media_thumbnail", None)
            if mthumbs and isinstance(mthumbs, list) and mthumbs:
                thumb = mthumbs[0].get("url", "")
            elif not thumb and "yt_videoid" in getattr(entry, "keys", lambda: [])():
                thumb = f"https://i.ytimg.com/vi/{entry['yt_videoid']}/hqdefault.jpg"
            score, matched = score_article(title, desc, keywords)
            if score > 0:
                articles.append(make_article(
                    title, link, f"YouTube · {channel_name}", "video",
                    date, score, matched, desc,
                    thumbnail=thumb,
                ))
    except Exception as e:
        error = str(e)
        print(f"    ERROR: {error}")

    record_stats("Tier 4", "YouTube (RSS)", channel_name, "video",
                 len(articles), entries_count, error,
                 notes=f"channel_id: {channel_id}")
    print(f"    -> {len(articles)} matched (of {entries_count} videos)")
    return articles


# ============================================================================
# FETCHER: Reddit
# ============================================================================

# Reddit blocks unauthenticated requests from cloud provider IPs (like
# Railway's). We use Client Credentials OAuth — register an app at
# reddit.com/prefs/apps and set REDDIT_CLIENT_ID + REDDIT_CLIENT_SECRET.
# Token is cached in-process so we authenticate once per scrape run.
_REDDIT_TOKEN_CACHE = {"token": None, "expires_at": 0}


def _get_reddit_token():
    """Fetch a Reddit OAuth token via client-credentials grant. Returns None if
    credentials aren't configured or the auth call fails."""
    client_id     = os.environ.get("REDDIT_CLIENT_ID")
    client_secret = os.environ.get("REDDIT_CLIENT_SECRET")
    if not client_id or not client_secret:
        return None

    now = time.time()
    if _REDDIT_TOKEN_CACHE["token"] and now < _REDDIT_TOKEN_CACHE["expires_at"]:
        return _REDDIT_TOKEN_CACHE["token"]

    try:
        r = requests.post(
            "https://www.reddit.com/api/v1/access_token",
            auth=(client_id, client_secret),
            data={"grant_type": "client_credentials"},
            headers={"User-Agent": REDDIT_UA},
            timeout=REQUEST_TIMEOUT,
        )
        if r.status_code != 200:
            print(f"    [reddit-oauth] token fetch HTTP {r.status_code}: {r.text[:160]}")
            return None
        data = r.json()
        token = data.get("access_token")
        expires_in = int(data.get("expires_in", 3600))
        if token:
            _REDDIT_TOKEN_CACHE["token"] = token
            _REDDIT_TOKEN_CACHE["expires_at"] = now + expires_in - 60  # refresh 60s early
        return token
    except Exception as e:
        print(f"    [reddit-oauth] token fetch error: {e}")
        return None


def fetch_reddit(subreddit, keywords):
    """Pull recent posts from a subreddit. Uses OAuth when REDDIT_CLIENT_ID +
    REDDIT_CLIENT_SECRET are set (needed on Railway since Reddit blocks
    cloud-IP unauth'd requests); falls back to the public JSON endpoint
    otherwise (works from most developer machines)."""
    print(f"  Reddit: r/{subreddit} ...")
    error = ""
    posts_seen = 0
    articles = []

    token = _get_reddit_token()
    if token:
        url = f"https://oauth.reddit.com/r/{subreddit}/new"
        headers = {"User-Agent": REDDIT_UA, "Authorization": f"Bearer {token}"}
    else:
        url = f"https://www.reddit.com/r/{subreddit}/new.json"
        headers = {"User-Agent": REDDIT_UA}

    try:
        r = requests.get(
            url, params={"limit": MAX_PER_CALL},
            timeout=REQUEST_TIMEOUT, headers=headers,
        )
        if r.status_code != 200:
            raise Exception(f"HTTP {r.status_code}")
        ctype = r.headers.get("Content-Type", "")
        if "application/json" not in ctype:
            raise Exception(f"Non-JSON response (content-type: {ctype[:60]}) — likely bot-blocked or rate-limited")
        children = r.json().get("data", {}).get("children", [])
        posts_seen = len(children)
        for child in children:
            post = child.get("data", {})
            title   = post.get("title", "")
            body    = post.get("selftext", "") or ""
            link    = post.get("url_overridden_by_dest") or f"https://reddit.com{post.get('permalink', '')}"
            created = post.get("created_utc")
            date    = datetime.fromtimestamp(created, tz=timezone.utc).isoformat() if created else ""
            score, matched = score_article(title, body, keywords)
            if score > 0:
                articles.append(make_article(title, link, f"Reddit · r/{subreddit}", "forums", date, score, matched, body))
    except Exception as e:
        error = str(e)
        print(f"    ERROR: {error}")

    notes = "OAuth (REDDIT_CLIENT_ID/SECRET set)" if token else "Unauthenticated (will fail on cloud IPs)"
    record_stats("Tier 3a", "Reddit", f"r/{subreddit}", "forums",
                 len(articles), posts_seen, error, notes=notes)
    print(f"    -> {len(articles)} matched (of {posts_seen} posts) [{notes}]")
    return articles


# ============================================================================
# FETCHER: Hacker News (uses ALL keywords, one Algolia query each)
# ============================================================================

def fetch_hackernews(keywords):
    """Search HN for each keyword. Algolia is fast and unrate-limited."""
    print("  HN: searching all keywords ...")
    all_articles = []
    error = ""
    total_seen = 0
    seen_ids = set()

    for kw in keywords:
        try:
            r = requests.get(
                "https://hn.algolia.com/api/v1/search_by_date",
                params={"query": kw, "tags": "story", "hitsPerPage": MAX_PER_CALL},
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            hits = r.json().get("hits", [])
        except Exception as e:
            error = str(e)
            print(f"    ERROR on '{kw}': {error}")
            continue
        total_seen += len(hits)
        for hit in hits:
            obj_id = hit.get("objectID", "")
            if obj_id in seen_ids:
                continue
            seen_ids.add(obj_id)
            title   = hit.get("title") or hit.get("story_title") or ""
            url_out = hit.get("url") or f"https://news.ycombinator.com/item?id={obj_id}"
            date    = hit.get("created_at", "")
            author  = hit.get("author", "")
            score, matched = score_article(title, "", keywords)
            if score > 0:
                all_articles.append(make_article(title, url_out, f"Hacker News · @{author}", "forums", date, score, matched, ""))
        time.sleep(API_DELAY / 2)  # HN is permissive; shorter delay

    record_stats("Tier 3a", "Hacker News (Algolia)",
                 "https://hn.algolia.com/api/v1/search_by_date",
                 "forums", len(all_articles), total_seen, error,
                 notes=f"Queried {len(keywords)} keywords separately")
    print(f"    -> {len(all_articles)} matched (of {total_seen} stories across {len(keywords)} queries)")
    return all_articles


# ============================================================================
# FETCHER: Stack Exchange (curated queries, not per-keyword, due to quota)
# ============================================================================

def fetch_stack_exchange(site, query, keywords):
    print(f"  StackExchange: {site} '{query}' ...")
    error = ""
    items_seen = 0
    articles = []
    try:
        r = requests.get(
            "https://api.stackexchange.com/2.3/search/advanced",
            params={"order": "desc", "sort": "creation", "site": site,
                    "q": query, "pagesize": MAX_PER_CALL},
            timeout=REQUEST_TIMEOUT, headers=HEADERS,
        )
        items = r.json().get("items", [])
        items_seen = len(items)
        for item in items:
            title   = item.get("title", "")
            link    = item.get("link", "")
            created = item.get("creation_date")
            date    = datetime.fromtimestamp(created, tz=timezone.utc).isoformat() if created else ""
            tags    = " ".join(item.get("tags", []))
            score, matched = score_article(title, tags, keywords)
            if score > 0:
                articles.append(make_article(title, link, f"Stack Exchange · {site}", "forums", date, score, matched, tags))
    except Exception as e:
        error = str(e)
        print(f"    ERROR: {error}")

    record_stats("Tier 3a", "Stack Exchange", f"{site}: {query}", "forums",
                 len(articles), items_seen, error,
                 notes="Curated query (SE has 300 req/day quota without API key)")
    print(f"    -> {len(articles)} matched (of {items_seen} questions)")
    return articles


# ============================================================================
# FETCHER: Bluesky (uses ALL keywords)
# ============================================================================

def fetch_bluesky(keywords):
    print("  Bluesky: searching all keywords ...")
    all_articles = []
    errors = []
    total_seen = 0
    seen_uris = set()

    for kw in keywords:
        try:
            # Note: public.api.bsky.app started returning 403 in late 2025;
            # api.bsky.app accepts anonymous requests fine.
            r = requests.get(
                "https://api.bsky.app/xrpc/app.bsky.feed.searchPosts",
                params={"q": kw, "limit": 25},
                timeout=REQUEST_TIMEOUT, headers=HEADERS,
            )
            posts = r.json().get("posts", [])
        except Exception as e:
            errors.append(f"{kw}: {e}")
            continue
        total_seen += len(posts)
        for post in posts:
            uri = post.get("uri", "")
            if uri in seen_uris:
                continue
            seen_uris.add(uri)
            text    = post.get("record", {}).get("text", "")
            author  = post.get("author", {}).get("handle", "")
            date    = post.get("record", {}).get("createdAt", "")
            if uri.startswith("at://"):
                parts = uri.replace("at://", "").split("/")
                if len(parts) >= 3 and author:
                    link = f"https://bsky.app/profile/{author}/post/{parts[2]}"
                else:
                    link = f"https://bsky.app/search?q={urllib.parse.quote(kw)}"
            else:
                link = f"https://bsky.app/search?q={urllib.parse.quote(kw)}"
            title = text[:80].replace("\n", " ").strip()
            if len(text) > 80: title += "…"
            if not title: continue
            score, matched = score_article(text, "", keywords)
            if score > 0:
                all_articles.append(make_article(title, link, f"Bluesky · @{author}", "social", date, score, matched, text))
        time.sleep(API_DELAY / 2)

    first_error = errors[0] if errors else ""
    record_stats("Tier 3b", "Bluesky",
                 "https://api.bsky.app/xrpc/app.bsky.feed.searchPosts",
                 "social", len(all_articles), total_seen, first_error,
                 notes=f"Queried {len(keywords)} keywords; {len(errors)} errored")
    print(f"    -> {len(all_articles)} matched (of {total_seen} posts; {len(errors)} query errors)")
    return all_articles


# ============================================================================
# LLM BLURB GENERATION
# ============================================================================
# For the articles most likely to reach the Media Director's eye (Breaking,
# Recent, and any direct-QB article anywhere in the feed), we generate a
# one-sentence plain-English "why it matters" using Claude Haiku.
# Blurbs are cached by link — re-runs cost $0 for articles we've seen.

BLURB_MODEL       = "claude-haiku-4-5"
BLURB_MAX_TOKENS  = 150
BLURB_CACHE_FILE  = os.path.join(DATA_DIR, "blurbs_cache.json")
# Coverage targets — generous enough that Breaking (1 per category), Recent,
# and any direct-QB article all get blurbs.
BLURB_TOP_OVERALL      = 20   # top N most recent overall
BLURB_TOP_NONDIRECT    = 20   # top N most recent non-direct (covers Recent Adjacent)
BLURB_TOP_PER_CATEGORY = 10   # top N per category (guarantees Breaking coverage)

# Direct-QB phrase matches (kept in sync with the list in index.html)
DIRECT_QB_TERMS = [
    "quantum biolog", "quantum bioscien", "quantum bio-",
    "biological quantum", "biology of quantum",
    "quantum effects in biolog", "quantum effects in living", "quantum effects in life",
    "quantum phenomena in biolog", "quantum phenomena in living", "quantum phenomena in life",
    "quantum mechanics in biolog", "quantum mechanics in living",
    "quantum mechanics of biolog", "quantum mechanics of life",
    "quantum processes in biolog", "quantum processes in living",
    "quantum behavior in biolog", "quantum behavior in living",
    "quantum life", "quantum living", "living quantum",
    "quantum nature of biolog", "quantum nature of life",
    "quantum-enabled biolog", "quantum enabled biolog",
    # Core mechanism — current flagship hypothesis of the field
    "radical pair",  # also catches "radical pair mechanism"
]

BLURB_SYSTEM_PROMPT = (
    "You are writing one- or two-sentence plain-English summaries of "
    "scientific articles for QUBIE News, a news aggregator for the "
    "Quantum Biology DAO.\n\n"
    "Your reader is an intelligent media director who is not a physicist "
    "or biochemist. They need to quickly understand (a) what the article "
    "is actually about, demystified, and (b) why it connects to quantum "
    "biology or an adjacent field (if not obvious).\n\n"
    "Rules:\n"
    "- Neutral and explanatory. No hype, no editorializing, no "
    "'this could revolutionize' language.\n"
    "- Do not start with 'This paper', 'The authors', 'This study' — "
    "just describe the finding directly.\n"
    "- If the article only tangentially relates to quantum biology "
    "(e.g. matched a keyword in passing), say so briefly.\n"
    "- Never invent details not supported by the title or summary.\n"
    "- Keep the whole response under 40 words.\n"
    "- No preamble, no sign-off — just the sentence(s)."
)


def is_direct_qb(article):
    """True if the article's title or summary mentions quantum biology explicitly."""
    text = ((article.get("title") or "") + " " + (article.get("summary") or "")).lower()
    return any(t in text for t in DIRECT_QB_TERMS)


def load_blurb_cache():
    if os.path.exists(BLURB_CACHE_FILE):
        try:
            with open(BLURB_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_blurb_cache(cache):
    with open(BLURB_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2, ensure_ascii=False)


def _generate_one_blurb(client, article):
    """Single API call for one article's blurb."""
    title = article.get("title", "").strip()
    summary = (article.get("summary") or "").strip()
    user_content = f"Title: {title}"
    if summary:
        user_content += f"\n\nSummary: {summary[:1000]}"
    resp = client.messages.create(
        model=BLURB_MODEL,
        max_tokens=BLURB_MAX_TOKENS,
        system=BLURB_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_content}],
    )
    return resp.content[0].text.strip()


def enrich_with_blurbs(articles):
    """Add a `blurb` field to qualifying articles. Uses cache aggressively."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key or api_key == "PASTE_YOUR_KEY_HERE":
        print("\n[Blurbs] No ANTHROPIC_API_KEY set in .env — skipping blurb generation.")
        return
    if anthropic is None:
        print("\n[Blurbs] anthropic package not installed — skipping blurb generation.")
        return

    print("\n[Blurbs] Generating plain-English summaries ...")

    client = anthropic.Anthropic(api_key=api_key)
    cache = load_blurb_cache()

    # Which articles qualify for blurbs — union of three buckets so that
    # Breaking, Recent·Q.B., and Recent·Adjacent Fields all get covered:
    #   1. Top N most recent overall   (Breaking has any-category content)
    #   2. All direct-QB articles      (Recent·Q.B. + any QB in the archive)
    #   3. Top N most recent non-direct (guarantees Recent·Adjacent coverage)
    articles_by_link = {a["link"]: a for a in articles if a.get("link")}
    to_blurb = set()
    # Bucket 1: overall top
    for a in articles[:BLURB_TOP_OVERALL]:
        if a.get("link"):
            to_blurb.add(a["link"])
    # Bucket 2: all direct-QB
    for a in articles:
        if a.get("link") and is_direct_qb(a):
            to_blurb.add(a["link"])
    # Bucket 3: non-direct top (Recent·Adjacent Fields)
    nondirect_recent = [a for a in articles if a.get("link") and not is_direct_qb(a)]
    for a in nondirect_recent[:BLURB_TOP_NONDIRECT]:
        to_blurb.add(a["link"])
    # Bucket 4: top-N per category (guarantees Breaking items always have blurbs)
    per_cat = {}
    for a in articles:
        cat = a.get("source_category")
        if cat and a.get("link"):
            per_cat.setdefault(cat, []).append(a)
    for cat, items in per_cat.items():
        for a in items[:BLURB_TOP_PER_CATEGORY]:
            to_blurb.add(a["link"])

    cached_hits = 0
    generated = 0
    errors = 0

    for link in to_blurb:
        article = articles_by_link.get(link)
        if not article:
            continue
        if link in cache:
            article["blurb"] = cache[link]
            cached_hits += 1
            continue
        try:
            blurb = _generate_one_blurb(client, article)
            article["blurb"] = blurb
            cache[link] = blurb
            generated += 1
            time.sleep(0.25)  # be polite, stay well under rate limits
        except Exception as e:
            errors += 1
            print(f"  blurb error: {e}")

    save_blurb_cache(cache)
    print(f"  -> {len(to_blurb)} articles targeted "
          f"({cached_hits} from cache, {generated} newly generated, {errors} errors)")


# ============================================================================
# XLSX WRITER
# ============================================================================

QBI_CREAM     = "EEE8DF"
QBI_PINK      = "F4CCCC"     # broken — soft red
QBI_GREEN     = "D9EAD3"     # working — soft green
QBI_YELLOW    = "FFF2CC"     # working-no-hits — soft yellow

ARIAL         = "Arial"
HEADER_FONT   = Font(name=ARIAL, bold=True, size=11)
BODY_FONT     = Font(name=ARIAL, size=10)
BOLD_BODY     = Font(name=ARIAL, size=10, bold=True)
ITALIC_SMALL  = Font(name=ARIAL, size=10, italic=True)
HEADER_FILL   = PatternFill("solid", start_color=QBI_CREAM)
BROKEN_FILL   = PatternFill("solid", start_color=QBI_PINK)
WORKING_FILL  = PatternFill("solid", start_color=QBI_GREEN)
NOHIT_FILL    = PatternFill("solid", start_color=QBI_YELLOW)


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)


def _apply_body_font(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.font = BODY_FONT


def _fill_for_status(status):
    if status == "ERROR":
        return BROKEN_FILL
    if status == "Working":
        return WORKING_FILL
    if status in ("Working (no hits)", "No results returned"):
        return NOHIT_FILL
    return None


REQUESTS_TAB_NAME    = "Source Requests"
KW_SUGGEST_TAB_NAME  = "Keyword Suggestions"
AUTO_TAB_NAMES = [
    "Summary",
    "Tier 1 - RSS Feeds",
    "Tier 2 - Search APIs",
    "Tier 3a - Forums",
    "Tier 3b - Social",
]
USER_TAB_NAMES = [REQUESTS_TAB_NAME, KW_SUGGEST_TAB_NAME]


def _ensure_requests_tab(wb):
    """Create the user-owned Source Requests tab if it doesn't exist yet."""
    if REQUESTS_TAB_NAME in wb.sheetnames:
        return  # already there — do nothing, preserve user's entries
    ws = wb.create_sheet(REQUESTS_TAB_NAME)
    headers = ["Source Name", "Type", "URL or Endpoint", "Why / Notes",
               "API Key Needed?", "Priority", "Status"]
    ws.append(headers)
    # Example row so the format is obvious
    example = ["e.g. Biophysical Journal", "RSS", "https://www.cell.com/biophysj/rss",
               "Core quantum biology / biophysics journal", "No", "High", "Requested"]
    ws.append(example)
    # Style header
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # Style example row — italic gray so it's obviously a template
    example_font = Font(name=ARIAL, size=10, italic=True, color="888888")
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).font = example_font
    ws.freeze_panes = "A2"
    _autofit(ws)


def write_sources_xlsx(path, unique_count):
    """Write live stats. Preserves the Source Requests tab across runs."""
    # Load existing workbook if present so we can preserve the user's Source
    # Requests tab. If not, start fresh.
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    # Remove any default "Sheet" from a fresh workbook
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Delete the auto-generated tabs so we can rebuild them with fresh stats.
    # (The Source Requests tab is intentionally NOT in this list, so it survives.)
    for name in AUTO_TAB_NAMES:
        if name in wb.sheetnames:
            del wb[name]

    # -- Summary tab --
    ws = wb.create_sheet("Summary")
    ws.append(["Tier", "Sources Configured", "Articles Contributed", "Errored"])
    tiers = {}
    for s in STATS_ROWS:
        t = s["tier"]
        if t not in tiers:
            tiers[t] = {"sources": 0, "matched": 0, "errored": 0}
        tiers[t]["sources"] += 1
        tiers[t]["matched"] += s["matched"]
        if s["status"] == "ERROR":
            tiers[t]["errored"] += 1
    for t in sorted(tiers.keys()):
        d = tiers[t]
        ws.append([t, d["sources"], d["matched"], d["errored"]])
    total_row = len(tiers) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row - 1})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row - 1})")
    ws.cell(row=total_row + 2, column=1, value=f"Last run: {datetime.now().isoformat(timespec='seconds')}")
    ws.cell(row=total_row + 3, column=1, value=f"Total unique articles after dedup: {unique_count}")
    _style_header(ws, 4)
    _apply_body_font(ws)
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).font = BOLD_BODY
    ws.cell(row=total_row + 2, column=1).font = ITALIC_SMALL
    ws.cell(row=total_row + 3, column=1).font = ITALIC_SMALL
    _autofit(ws)

    # -- One tab per tier --
    tier_order = ["Tier 1", "Tier 2", "Tier 3a", "Tier 3b"]
    tier_titles = {
        "Tier 1":  "Tier 1 - RSS Feeds",
        "Tier 2":  "Tier 2 - Search APIs",
        "Tier 3a": "Tier 3a - Forums",
        "Tier 3b": "Tier 3b - Social",
    }
    for t in tier_order:
        ws = wb.create_sheet(tier_titles[t])
        headers = ["Source Name", "Target / URL", "Category", "Articles Matched",
                   "Total Returned", "Status", "Error", "Notes"]
        ws.append(headers)
        rows = [s for s in STATS_ROWS if s["tier"] == t]
        for s in rows:
            ws.append([s["source_name"], s["target"], s["category"],
                       s["matched"], s["total"], s["status"],
                       s["error"], s["notes"]])
        # Color rows by status
        for i, s in enumerate(rows, start=2):
            fill = _fill_for_status(s["status"])
            if fill:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=i, column=c).fill = fill
        _style_header(ws, len(headers))
        _apply_body_font(ws)
        _autofit(ws)

    # Create the user-owned Source Requests tab if it doesn't exist yet.
    _ensure_requests_tab(wb)

    # Force the intended tab order every run.
    # Auto tabs first, then user tabs (Source Requests, Keyword Suggestions).
    # Any other sheets that exist (e.g. created by the web server) come after.
    desired_order = AUTO_TAB_NAMES + USER_TAB_NAMES
    remaining = [name for name in wb.sheetnames if name not in desired_order]
    wb._sheets = (
        [wb[name] for name in desired_order if name in wb.sheetnames]
        + [wb[name] for name in remaining]
    )

    wb.save(path)


def write_sources_json(path, unique_count):
    """Export STATS_ROWS as JSON for fast browser loading on the /sources page."""
    by_tier = {}
    for s in STATS_ROWS:
        by_tier.setdefault(s["tier"], []).append(s)

    # Summary: one row per tier with aggregate counts
    summary = []
    for tier_name in sorted(by_tier.keys()):
        rows = by_tier[tier_name]
        summary.append({
            "tier": tier_name,
            "sources_configured": len(rows),
            "articles_contributed": sum(r["matched"] for r in rows),
            "errored": sum(1 for r in rows if r["status"] == "ERROR"),
        })

    output = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "unique_articles_after_dedup": unique_count,
        "summary": summary,
        "tiers": by_tier,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("QUBIE NEWS Scraper")
    print("=" * 60)
    print()

    keywords = load_keywords()
    print()

    all_articles = []

    # -- Tier 1: RSS Feeds --
    print("[Tier 1] RSS Feeds")
    for name, url, category in RSS_FEEDS:
        all_articles.extend(fetch_rss(name, url, category, keywords))
        time.sleep(API_DELAY)
    print()

    # -- Tier 2: Search APIs (chunked, uses ALL keywords) --
    print("[Tier 2] Search APIs")
    all_articles.extend(fetch_pubmed(keywords))
    all_articles.extend(fetch_arxiv_api(keywords))
    all_articles.extend(fetch_europepmc(keywords))
    all_articles.extend(fetch_semantic_scholar(keywords))
    all_articles.extend(fetch_openalex(keywords))
    print()

    # -- Tier 3a: Forums --
    print("[Tier 3a] Forums")
    for sub in REDDIT_SUBREDDITS:
        all_articles.extend(fetch_reddit(sub, keywords))
        time.sleep(API_DELAY)
    all_articles.extend(fetch_hackernews(keywords))
    for site, q in STACK_EXCHANGE:
        all_articles.extend(fetch_stack_exchange(site, q, keywords))
        time.sleep(API_DELAY)
    print()

    # -- Tier 3b: Social --
    print("[Tier 3b] Social")
    all_articles.extend(fetch_bluesky(keywords))
    print()

    # -- Tier 4: Video --
    print("[Tier 4] Video")
    all_articles.extend(fetch_youtube_api(keywords))
    for channel_name, channel_id in YOUTUBE_CHANNELS:
        all_articles.extend(fetch_youtube_channel(channel_name, channel_id, keywords))
        time.sleep(API_DELAY)
    print()

    # -- Deduplicate by title --
    seen = set()
    unique = []
    for a in all_articles:
        key = a["title"].lower().strip()
        if key and key not in seen:
            seen.add(key)
            unique.append(a)

    # -- Sort by date_iso desc (most recent first) --
    unique.sort(key=lambda a: a["date_iso"] or "0", reverse=True)

    # -- Enrich qualifying articles with LLM-generated blurbs --
    enrich_with_blurbs(unique)

    # -- Write feed.json --
    output = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "article_count": len(unique),
        "articles": unique,
    }
    out_path = os.path.abspath(OUTPUT_FILE)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # -- Write sources xlsx + JSON export (JSON for fast browser loading) --
    xlsx_path = os.path.abspath(XLSX_FILE)
    json_path = os.path.abspath(SOURCES_JSON)
    write_sources_xlsx(xlsx_path, len(unique))
    write_sources_json(json_path, len(unique))

    # -- Summary --
    by_category = {}
    for a in unique:
        by_category[a["source_category"]] = by_category.get(a["source_category"], 0) + 1

    print("=" * 60)
    print(f"Done! {len(unique)} unique articles (from {len(all_articles)} matches)")
    print(f"  By category: {by_category}")
    print(f"  Wrote: {out_path}")
    print(f"  Wrote: {xlsx_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
