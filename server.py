"""
QUBIE News local server
========================
Replaces `python -m http.server` with a Flask app that:
  - Serves the static site (index.html, feed.json, assets, etc.)
  - Provides an admin API so the Sources and Keywords pages can
    read/write the underlying config files.

Run:
    python server.py

Then visit:
    http://localhost:8000            -> the Report
    http://localhost:8000/sources    -> the Sources admin page
    http://localhost:8000/keywords   -> the Keywords admin page
"""
import json
import os
import shutil
import subprocess
import sys
import threading
from datetime import datetime
from functools import wraps

from flask import Flask, Response, jsonify, redirect, request, send_from_directory, session, abort
from openpyxl import load_workbook

import db  # SQLite layer for users + saves

# --- Paths ------------------------------------------------------------------
# Locally: files live next to server.py
# On Railway: DATA_DIR points to the persistent volume so user edits + scrape
# output survive redeploys.
HERE          = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.environ.get("DATA_DIR") or HERE
BUNDLED_KEYWORDS = os.path.join(HERE, "scraper", "keywords.txt")

KEYWORDS_FILE = os.path.join(DATA_DIR, "keywords.txt")
XLSX_FILE     = os.path.join(DATA_DIR, "QBIO-Report-Sources.xlsx")
SOURCES_JSON  = os.path.join(DATA_DIR, "sources.json")
FEED_JSON     = os.path.join(DATA_DIR, "feed.json")
LOG_FILE      = os.path.join(DATA_DIR, "source_requests_log.md")
REQUESTS_TAB  = "Source Requests"

# First-run seed: if the volume is empty, copy the bundled keywords.txt in so
# the scraper has a starting point. User edits thereafter stay on the volume.
def _seed_volume():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(KEYWORDS_FILE) and os.path.exists(BUNDLED_KEYWORDS):
        shutil.copyfile(BUNDLED_KEYWORDS, KEYWORDS_FILE)
        print(f"[seed] Copied bundled keywords.txt -> {KEYWORDS_FILE}")

_seed_volume()

# Initialize the SQLite DB (users + saves). Idempotent.
db.init_db()
print(f"[db] qubie.db ready at {db.DB_PATH}")
if db.BOOTSTRAP_ADMIN_USERNAME:
    print(f"[db] bootstrap admin username: '{db.BOOTSTRAP_ADMIN_USERNAME}' "
          f"(first matching signup will be promoted)")

# ============================================================================
# Auth — cookie-based session with a password-only login form.
# Admin pages redirect unauthed users to /login; admin APIs return 401 JSON.
# ============================================================================

# Admin password for /admin. MUST be set as an env var in production.
# The placeholder below only exists so the server can start locally for
# development without env vars; it is NOT a valid production password.
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD") or "please-set-ADMIN_PASSWORD-env-var"
# Note: app.secret_key is set a few lines below, right after app is created.


def _current_user():
    """Return the current logged-in user dict, or None.
    Honors both the new user_id session and the legacy 'authed' flag (the
    legacy flag means the bearer typed the ADMIN_PASSWORD; treat as admin)."""
    uid = session.get("user_id")
    if uid:
        u = db.get_user_by_id(uid)
        if u:
            return u
        # Stale session referencing a deleted user — clear it
        session.pop("user_id", None)
    if session.get("authed"):
        # Legacy ADMIN_PASSWORD path — synthesize a transient "admin" identity
        # so existing code that just needs is_admin keeps working.
        return {"id": 0, "username": "_legacy_admin", "is_admin": True,
                "bio": "", "avatar_url": "", "created_at": "",
                "_legacy": True}
    return None


def is_admin() -> bool:
    u = _current_user()
    return bool(u and u.get("is_admin"))


def require_auth(f):
    """Allow either: new logged-in user OR legacy ADMIN_PASSWORD session.
    Existing /admin pages route through here — those pages additionally check
    is_admin() to gate admin-only operations."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if _current_user():
            return f(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "Login required"}), 401
        return redirect(f"/login?next={request.path}")
    return decorated


def require_admin(f):
    """Stricter than require_auth — must be is_admin."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if is_admin():
            return f(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "Admin access required"}), 403
        return redirect(f"/login?next={request.path}")
    return decorated

LOG_HEADER = (
    "# QUBIE News - Source Requests Log\n\n"
    "_Newest at top. Each entry is formatted for copy-paste straight to Claude._\n\n"
    "<!-- NEW_ENTRIES_BELOW -->\n\n"
)
LOG_SENTINEL = "<!-- NEW_ENTRIES_BELOW -->\n\n"

PORT = int(os.environ.get("PORT", 8000))

app = Flask(__name__, static_folder=HERE, static_url_path="")
# Session cookies are signed with this secret. MUST be set as an env var in
# production — the placeholder fallback is only for local dev so the server
# can boot without env vars. Rotating FLASK_SECRET_KEY invalidates all sessions.
app.secret_key = os.environ.get("FLASK_SECRET_KEY") or "please-set-FLASK_SECRET_KEY-env-var"


# ============================================================================
# Page routes
# ============================================================================

@app.route("/")
def page_report():
    return send_from_directory(HERE, "index.html")


@app.route("/login", methods=["GET", "POST"])
def page_login():
    """Login form. Accepts username + password (new) OR legacy
    ADMIN_PASSWORD-only (kept active during the transition so Olli is never
    locked out — to be removed in a follow-up commit once the new auth is
    confirmed working)."""
    nxt = request.args.get("next") or request.form.get("next") or "/"
    if not (isinstance(nxt, str) and nxt.startswith("/")):
        nxt = "/"

    error = ""
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        pw       = (request.form.get("password") or "").strip()

        if username:
            # New path: username + password
            user = db.authenticate(username, pw)
            if user:
                session.clear()
                session["user_id"] = user["id"]
                session.permanent = True
                return redirect(nxt)
            error = "Wrong username or password."
        else:
            # Legacy path: ADMIN_PASSWORD only (no username field).
            # Kept until new auth is verified; will be removed.
            if pw and pw == ADMIN_PASSWORD:
                session.clear()
                session["authed"] = True
                session.permanent = True
                return redirect(nxt)
            error = "Wrong password."

    with open(os.path.join(HERE, "login.html"), "r", encoding="utf-8") as f:
        html = f.read()
    next_query = f'?next={nxt}' if nxt and nxt != "/" else ""
    html = html.replace("{{NEXT_QUERY}}", next_query)
    html = html.replace("{{ERROR_MESSAGE}}", error)
    return html


@app.route("/signup", methods=["GET", "POST"])
def page_signup():
    """Open signup. Username + password. The first user matching the
    BOOTSTRAP_ADMIN_USERNAME env var is auto-promoted to admin."""
    nxt = request.args.get("next") or request.form.get("next") or "/"
    if not (isinstance(nxt, str) and nxt.startswith("/")):
        nxt = "/"

    error = ""
    if request.method == "POST":
        username  = (request.form.get("username") or "").strip()
        password  = (request.form.get("password") or "").strip()
        password2 = (request.form.get("password2") or "").strip()

        err = db.validate_username(username)
        if err: error = err
        if not error:
            err = db.validate_password(password)
            if err: error = err
        if not error and password != password2:
            error = "Passwords don't match."

        if not error:
            try:
                user = db.create_user(username, password)
                session.clear()
                session["user_id"] = user["id"]
                session.permanent = True
                return redirect(nxt)
            except ValueError as e:
                error = str(e)

    with open(os.path.join(HERE, "signup.html"), "r", encoding="utf-8") as f:
        html = f.read()
    next_query = f'?next={nxt}' if nxt and nxt != "/" else ""
    html = html.replace("{{NEXT_QUERY}}", next_query)
    html = html.replace("{{ERROR_MESSAGE}}", error)
    return html


@app.route("/logout")
def page_logout():
    session.clear()
    return redirect("/")


# ============================================================================
# User profile pages + APIs
# ============================================================================

@app.route("/me")
def page_me():
    user = _current_user()
    if not user or user.get("_legacy"):
        return redirect("/login?next=/me")
    return redirect(f"/u/{user['username']}")


@app.route("/u/<username>")
def page_profile(username):
    return send_from_directory(HERE, "profile.html")


@app.route("/api/me", methods=["GET"])
def api_me():
    """Return the current user's public profile, or {logged_in: False}."""
    user = _current_user()
    if not user:
        return jsonify({"logged_in": False})
    if user.get("_legacy"):
        # Legacy admin session — no real user account.
        return jsonify({
            "logged_in": True,
            "legacy_admin": True,
            "username":  "_legacy_admin",
            "is_admin":  True,
        })
    return jsonify({
        "logged_in": True,
        "id":         user["id"],
        "username":   user["username"],
        "bio":        user["bio"],
        "avatar_url": user["avatar_url"],
        "is_admin":   user["is_admin"],
        "created_at": user["created_at"],
    })


@app.route("/api/users/<username>", methods=["GET"])
def api_user_profile(username):
    """Public profile lookup."""
    user = db.get_user_by_username(username)
    if not user:
        return jsonify({"error": "User not found"}), 404
    saves = db.list_user_saves(user["id"])
    return jsonify({
        "username":   user["username"],
        "bio":        user["bio"],
        "avatar_url": user["avatar_url"],
        "is_admin":   user["is_admin"],
        "created_at": user["created_at"],
        "save_count": len(saves),
        "saves":      saves,
    })


@app.route("/api/me/profile", methods=["POST"])
def api_update_profile():
    """Update the current user's bio/avatar_url."""
    user = _current_user()
    if not user or user.get("_legacy"):
        return jsonify({"error": "Login required"}), 401
    data = request.get_json(force=True, silent=True) or {}
    bio        = (data.get("bio") or "").strip()[:500]
    avatar_url = (data.get("avatar_url") or "").strip()[:500]
    # Light sanity-check on avatar_url
    if avatar_url and not (avatar_url.startswith("http://") or avatar_url.startswith("https://")):
        return jsonify({"error": "Avatar URL must start with http:// or https://"}), 400
    db.update_profile(user["id"], bio=bio, avatar_url=avatar_url)
    return jsonify({"ok": True})


# ============================================================================
# Saves API
# ============================================================================

@app.route("/api/saves", methods=["POST"])
def api_save_article():
    """Save an article snapshot for the current user. Idempotent — saving
    the same link twice returns the existing save id."""
    user = _current_user()
    if not user or user.get("_legacy"):
        return jsonify({"error": "Login required"}), 401
    data = request.get_json(force=True, silent=True) or {}
    if not (data.get("link") or "").strip():
        return jsonify({"error": "link is required"}), 400
    save_id = db.add_save(user["id"], data)
    return jsonify({"ok": True, "save_id": save_id})


@app.route("/api/saves", methods=["DELETE"])
def api_unsave_article():
    """Unsave by article link. Body: {link: ...}"""
    user = _current_user()
    if not user or user.get("_legacy"):
        return jsonify({"error": "Login required"}), 401
    data = request.get_json(force=True, silent=True) or {}
    link = (data.get("link") or "").strip()
    if not link:
        return jsonify({"error": "link is required"}), 400
    removed = db.remove_save(user["id"], link)
    return jsonify({"ok": True, "removed": removed})


@app.route("/api/saves/me", methods=["GET"])
def api_my_saved_links():
    """Return just the set of links the current user has saved — used by
    the bookmark icons to know which are filled vs hollow."""
    user = _current_user()
    if not user or user.get("_legacy"):
        return jsonify({"logged_in": False, "links": []})
    return jsonify({
        "logged_in": True,
        "links":     list(db.list_user_saved_links(user["id"])),
    })


@app.route("/api/saves/most", methods=["GET"])
def api_most_saved():
    """Top saved articles in the last 30 days, with last-saver attribution."""
    try:
        window = int(request.args.get("window") or 30)
    except (TypeError, ValueError):
        window = 30
    try:
        limit = int(request.args.get("limit") or 10)
    except (TypeError, ValueError):
        limit = 10
    return jsonify({
        "window_days": window,
        "items":       db.most_saved(window_days=window, limit=limit),
    })


@app.route("/chatter")
def page_chatter():
    return send_from_directory(HERE, "chatter.html")


@app.route("/video")
def page_video():
    return send_from_directory(HERE, "video.html")


@app.route("/suggest")
def page_suggest():
    """Public suggestion page for DAO members."""
    return send_from_directory(HERE, "suggest.html")


@app.route("/admin")
@require_auth
def page_admin():
    return send_from_directory(HERE, "admin.html")


# Old admin URLs redirect to the consolidated /admin page (bookmarks still work)
@app.route("/sources")
def page_sources_redirect():
    return redirect("/admin", code=302)


@app.route("/keywords")
def page_keywords_redirect():
    return redirect("/admin", code=302)


# These are data files on the volume (not static assets) — serve from DATA_DIR
@app.route("/feed.json")
def serve_feed():
    if not os.path.exists(FEED_JSON):
        return jsonify({"error": "No feed yet — scraper hasn't run."}), 404
    return send_from_directory(DATA_DIR, "feed.json")


SCRAPE_PROGRESS_FILE = os.path.join(DATA_DIR, "scrape_progress.json")


@app.route("/api/scrape-progress", methods=["GET"])
def api_scrape_progress():
    """Return the current scrape progress (written by scraper after each source).
    No-auth — polled by /admin during a scrape. Returns {running, done, total,
    percent, current, started_at, finished_at} or {running: false} if no file."""
    if not os.path.exists(SCRAPE_PROGRESS_FILE):
        return jsonify({"running": False})
    try:
        with open(SCRAPE_PROGRESS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
    except (json.JSONDecodeError, OSError):
        return jsonify({"running": False})
    total   = int(data.get("total") or 0)
    done    = int(data.get("done") or 0)
    percent = int(round(100 * done / total)) if total else 0
    finished = bool(data.get("finished_at"))
    return jsonify({
        "running":      not finished,
        "done":         done,
        "total":        total,
        "percent":      percent,
        "current":      data.get("current") or "",
        "started_at":   data.get("started_at"),
        "finished_at": data.get("finished_at"),
    })


@app.route("/sources.json")
def serve_sources_json():
    if not os.path.exists(SOURCES_JSON):
        return jsonify({"error": "No sources.json yet — scraper hasn't run."}), 404
    return send_from_directory(DATA_DIR, "sources.json")


# ============================================================================
# Keywords API
# ============================================================================

def _read_keywords_file():
    """Return (lines, keywords_only) where lines is the raw file and
    keywords_only is the non-comment, non-blank entries."""
    if not os.path.exists(KEYWORDS_FILE):
        return [], []
    with open(KEYWORDS_FILE, "r", encoding="utf-8") as f:
        lines = f.read().splitlines()
    keywords = [
        ln.strip() for ln in lines
        if ln.strip() and not ln.strip().startswith("#")
    ]
    return lines, keywords


def _write_keywords_lines(lines):
    with open(KEYWORDS_FILE, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


@app.route("/api/keywords", methods=["GET"])
def api_list_keywords():
    """Public read: DAO members can see current keywords (including weights)."""
    lines, keywords = _read_keywords_file()
    return jsonify({"keywords": keywords, "count": len(keywords)})


import re as _re
_KW_WEIGHT_RE = _re.compile(r"^(.*?)\s*\[(\d+)\]\s*$")


def _extract_phrase(s):
    """Return just the phrase (lowercased) from a possibly-weighted entry."""
    s = (s or "").strip()
    m = _KW_WEIGHT_RE.match(s)
    return (m.group(1) if m else s).strip().lower()


@app.route("/api/keywords", methods=["POST"])
@require_auth
def api_add_keyword():
    data = request.get_json(force=True, silent=True) or {}
    new_kw = (data.get("keyword") or "").strip()
    if not new_kw:
        return jsonify({"error": "Empty keyword"}), 400
    if new_kw.startswith("#"):
        return jsonify({"error": "Keywords cannot start with '#'"}), 400

    new_phrase = _extract_phrase(new_kw)
    if not new_phrase:
        return jsonify({"error": "Empty phrase"}), 400

    lines, _ = _read_keywords_file()
    # Remove any existing line with the same phrase (regardless of weight).
    # This makes "adding" a weighted version replace the unweighted one in place.
    replaced = False
    kept = []
    for ln in lines:
        stripped = ln.strip()
        if not stripped or stripped.startswith("#"):
            kept.append(ln)
            continue
        if _extract_phrase(stripped) == new_phrase:
            replaced = True
            continue  # drop this line
        kept.append(ln)

    # Append the new entry at the end
    if kept and kept[-1].strip():
        kept.append("")
    kept.append(new_kw)
    _write_keywords_lines(kept)
    return jsonify({"ok": True, "keyword": new_kw, "replaced": replaced})


@app.route("/api/keywords/<path:keyword>", methods=["DELETE"])
@require_auth
def api_delete_keyword(keyword):
    target = keyword.strip().lower()
    lines, _ = _read_keywords_file()
    new_lines = [ln for ln in lines if ln.strip().lower() != target]
    if len(new_lines) == len(lines):
        return jsonify({"error": "Not found"}), 404
    _write_keywords_lines(new_lines)
    return jsonify({"ok": True, "keyword": keyword})


# ============================================================================
# Source Requests API (reads/writes Tab 6 of the xlsx)
# ============================================================================

REQUEST_HEADERS = ["Source Name", "Type", "URL or Endpoint", "Why / Notes",
                   "API Key Needed?", "Priority", "Status"]


def _is_example_row(row):
    """Is this the italic-gray template row we insert on first run?"""
    return row and isinstance(row[0], str) and row[0].startswith("e.g.")


def _read_source_requests():
    if not os.path.exists(XLSX_FILE):
        return []
    wb = load_workbook(XLSX_FILE)
    if REQUESTS_TAB not in wb.sheetnames:
        return []
    ws = wb[REQUESTS_TAB]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c not in (None, "") for c in row):
            continue
        if _is_example_row(row):
            continue
        rows.append({
            "row_number":      i,
            "source_name":     row[0] or "",
            "type":            row[1] or "",
            "url":             row[2] or "",
            "why_notes":       row[3] or "",
            "api_key_needed":  row[4] or "",
            "priority":        row[5] or "",
            "status":          row[6] or "Requested",
        })
    return rows


@app.route("/api/source-requests", methods=["GET"])
@require_auth
def api_list_source_requests():
    return jsonify({"requests": _read_source_requests()})


def _log_source_request(payload):
    """Prepend a new source request entry to source_requests_log.md.
    Newest-at-top so you skim the latest right away."""
    timestamp = datetime.now().isoformat(timespec="seconds")
    entry = (
        f"## {timestamp}\n\n"
        f"- **Source:** {payload.get('source_name', '')}\n"
        f"- **Type:** {payload.get('type', '')}\n"
        f"- **URL / Target:** {payload.get('url', '')}\n"
        f"- **Why / Notes:** {payload.get('why_notes', '')}\n"
        f"- **API Key Needed:** {payload.get('api_key_needed', 'Unknown')}\n"
        f"- **Priority:** {payload.get('priority', 'Medium')}\n"
        f"- **Status:** {payload.get('status', 'Requested')}\n\n"
        f"---\n\n"
    )
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, "w", encoding="utf-8") as f:
            f.write(LOG_HEADER)
    with open(LOG_FILE, "r", encoding="utf-8") as f:
        content = f.read()
    if LOG_SENTINEL in content:
        content = content.replace(LOG_SENTINEL, LOG_SENTINEL + entry, 1)
    else:
        content = LOG_HEADER + entry + content
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write(content)


@app.route("/api/source-requests", methods=["POST"])
@require_auth
def api_add_source_request():
    data = request.get_json(force=True, silent=True) or {}
    if not (data.get("source_name") or "").strip():
        return jsonify({"error": "source_name is required"}), 400

    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "Sources workbook not found. Run the scraper first."}), 400

    wb = load_workbook(XLSX_FILE)
    if REQUESTS_TAB not in wb.sheetnames:
        return jsonify({"error": f"Tab '{REQUESTS_TAB}' missing. Run scraper first."}), 500
    ws = wb[REQUESTS_TAB]
    payload = {
        "source_name":     data.get("source_name", "").strip(),
        "type":            data.get("type", "").strip(),
        "url":             data.get("url", "").strip(),
        "why_notes":       data.get("why_notes", "").strip(),
        "api_key_needed":  data.get("api_key_needed", "").strip() or "Unknown",
        "priority":        data.get("priority", "").strip() or "Medium",
        "status":          data.get("status", "").strip() or "Requested",
    }
    ws.append([payload["source_name"], payload["type"], payload["url"],
               payload["why_notes"], payload["api_key_needed"],
               payload["priority"], payload["status"]])
    wb.save(XLSX_FILE)

    # Also prepend to the copy-paste-ready log file
    try:
        _log_source_request(payload)
    except Exception as e:
        # Log failure shouldn't block the save — just report it
        print(f"WARN: failed to write source_requests_log.md: {e}")

    return jsonify({"ok": True, "logged_to": "source_requests_log.md"})


# ----------------------------------------------------------------------------
# Runtime sources config — where /admin's "Push" button writes
# ----------------------------------------------------------------------------
SOURCES_CONFIG_FILE = os.path.join(DATA_DIR, "sources_config.json")


def _load_sources_config():
    if not os.path.exists(SOURCES_CONFIG_FILE):
        return {"rss_feeds": [], "reddit_subreddits": []}
    try:
        with open(SOURCES_CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except (json.JSONDecodeError, OSError):
        cfg = {}
    cfg.setdefault("rss_feeds", [])
    cfg.setdefault("reddit_subreddits", [])
    return cfg


def _save_sources_config(cfg):
    with open(SOURCES_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


_SUBREDDIT_URL_RE = _re.compile(r"reddit\.com/r/([A-Za-z0-9_]+)", _re.IGNORECASE)


def _extract_subreddit(name_or_url):
    """Pull a subreddit name from a URL or a bare 'r/foo' / 'foo' string."""
    s = (name_or_url or "").strip()
    if not s:
        return None
    m = _SUBREDDIT_URL_RE.search(s)
    if m:
        return m.group(1)
    # bare "r/foo" or "/r/foo"
    s_stripped = s.lstrip("/").lstrip("r/").lstrip()
    if s_stripped and "/" not in s_stripped and " " not in s_stripped and "." not in s_stripped:
        return s_stripped
    return None


# ----------------------------------------------------------------------------
# Public suggestion endpoints (no auth) — for DAO members via /suggest
# ----------------------------------------------------------------------------
KW_SUGGEST_TAB = "Keyword Suggestions"


def _ensure_kw_suggest_tab(wb):
    """Create Keyword Suggestions tab if it doesn't exist yet."""
    if KW_SUGGEST_TAB in wb.sheetnames:
        return
    ws = wb.create_sheet(KW_SUGGEST_TAB)
    headers = ["Suggested Keyword", "Suggested Weight", "Why / Notes",
               "Submitted By", "Submitted At", "Status"]
    ws.append(headers)
    ws.append(["e.g. microtubule coherence", "3",
               "Growing relevance for Penrose-Hameroff work",
               "jane@example.com", "2026-04-21T12:00:00", "Pending"])


@app.route("/api/suggest-source", methods=["POST"])
def api_suggest_source():
    """Public: DAO member suggests a new source. Flows into Source Requests tab."""
    data = request.get_json(force=True, silent=True) or {}
    source_name = (data.get("source_name") or "").strip()
    if not source_name:
        return jsonify({"error": "Source name is required"}), 400

    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "System not ready (no sources workbook yet)"}), 503

    submitted_by = (data.get("submitted_by") or "").strip() or "anonymous"
    why_notes = (data.get("why_notes") or "").strip()
    # Prefix the notes so Olli sees it came from the public form
    prefixed_notes = f"[Public suggestion by {submitted_by}] {why_notes}".strip()

    wb = load_workbook(XLSX_FILE)
    if REQUESTS_TAB not in wb.sheetnames:
        return jsonify({"error": "Requests tab missing"}), 500
    ws = wb[REQUESTS_TAB]
    payload = {
        "source_name":     source_name,
        "type":            (data.get("type") or "").strip(),
        "url":             (data.get("url") or "").strip(),
        "why_notes":       prefixed_notes,
        "api_key_needed":  (data.get("api_key_needed") or "").strip() or "Unknown",
        "priority":        (data.get("priority") or "").strip() or "Medium",
        "status":          "Requested (public)",
    }
    ws.append([payload["source_name"], payload["type"], payload["url"],
               payload["why_notes"], payload["api_key_needed"],
               payload["priority"], payload["status"]])
    wb.save(XLSX_FILE)

    # Also append to the markdown log (same as MD submissions)
    try:
        _log_source_request(payload)
    except Exception as e:
        print(f"WARN: failed to write source_requests_log.md: {e}")

    return jsonify({"ok": True})


@app.route("/api/suggest-keyword", methods=["POST"])
def api_suggest_keyword():
    """Public: DAO member suggests a new keyword OR requests a weight change
    on an existing one. Goes into Keyword Suggestions tab.

    If `kind` is "weight-change", the `why_notes` is prefixed with a tag so the
    admin UI can display it distinctly.
    """
    data = request.get_json(force=True, silent=True) or {}
    phrase = (data.get("phrase") or "").strip()
    if not phrase:
        return jsonify({"error": "Keyword is required"}), 400
    if phrase.startswith("#"):
        return jsonify({"error": "Keyword cannot start with '#'"}), 400

    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "System not ready (no sources workbook yet)"}), 503

    weight = 1
    try:
        weight = max(1, min(10, int(data.get("weight") or 1)))
    except (TypeError, ValueError):
        weight = 1

    why_notes = (data.get("why_notes") or "").strip()
    submitted_by = (data.get("submitted_by") or "").strip() or "anonymous"
    submitted_at = datetime.now().isoformat(timespec="seconds")

    # Tag weight-change requests so the admin UI can render them as "change"
    # rather than "new". Keeps the same tab/column shape.
    if (data.get("kind") or "").lower() == "weight-change":
        tag = "[weight-change]"
        why_notes = f"{tag} {why_notes}".strip()

    wb = load_workbook(XLSX_FILE)
    _ensure_kw_suggest_tab(wb)
    ws = wb[KW_SUGGEST_TAB]
    ws.append([phrase, weight, why_notes, submitted_by, submitted_at, "Pending"])
    wb.save(XLSX_FILE)
    return jsonify({"ok": True})


# --- Admin: review & action the keyword-suggestions queue ---

def _is_example_kw_suggest_row(row):
    return row and isinstance(row[0], str) and row[0].startswith("e.g.")


def _read_kw_suggestions():
    if not os.path.exists(XLSX_FILE):
        return []
    wb = load_workbook(XLSX_FILE)
    if KW_SUGGEST_TAB not in wb.sheetnames:
        return []
    ws = wb[KW_SUGGEST_TAB]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c not in (None, "") for c in row):
            continue
        if _is_example_kw_suggest_row(row):
            continue
        phrase = row[0] or ""
        weight = row[1] or 1
        notes  = row[2] or ""
        is_weight_change = str(notes).lstrip().lower().startswith("[weight-change]")
        # Strip the tag from display
        if is_weight_change:
            notes = _re.sub(r"^\s*\[weight-change\]\s*", "", str(notes), flags=_re.IGNORECASE)
        rows.append({
            "row_number":       i,
            "phrase":           phrase,
            "suggested_weight": int(weight) if str(weight).isdigit() else weight,
            "why_notes":        notes,
            "submitted_by":     row[3] or "",
            "submitted_at":     row[4] or "",
            "status":           row[5] or "Pending",
            "is_weight_change": is_weight_change,
        })
    return rows


@app.route("/api/keyword-suggestions", methods=["GET"])
@require_auth
def api_list_kw_suggestions():
    return jsonify({"suggestions": _read_kw_suggestions()})


@app.route("/api/keyword-suggestions/<int:row_number>/approve", methods=["POST"])
@require_auth
def api_approve_kw_suggestion(row_number):
    """Apply a pending keyword suggestion to keywords.txt, then remove the row."""
    if row_number < 2:
        return jsonify({"error": "Cannot approve header row"}), 400
    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "Workbook missing"}), 404

    wb = load_workbook(XLSX_FILE)
    if KW_SUGGEST_TAB not in wb.sheetnames:
        return jsonify({"error": "Suggestions tab missing"}), 404
    ws = wb[KW_SUGGEST_TAB]
    if row_number > ws.max_row:
        return jsonify({"error": "Row out of range"}), 404

    phrase = (ws.cell(row=row_number, column=1).value or "").strip()
    weight_raw = ws.cell(row=row_number, column=2).value
    try:
        weight = max(1, min(10, int(weight_raw)))
    except (TypeError, ValueError):
        weight = 1

    if not phrase:
        return jsonify({"error": "Empty phrase in this row"}), 400

    # Apply to keywords.txt by reusing the add-keyword replacement logic.
    new_kw = phrase if weight == 1 else f"{phrase} [{weight}]"
    new_phrase = _extract_phrase(new_kw)
    lines, _ = _read_keywords_file()
    kept = []
    for ln in lines:
        stripped = ln.strip()
        if not stripped or stripped.startswith("#"):
            kept.append(ln)
            continue
        if _extract_phrase(stripped) == new_phrase:
            continue  # drop — will be replaced
        kept.append(ln)
    if kept and kept[-1].strip():
        kept.append("")
    kept.append(new_kw)
    _write_keywords_lines(kept)

    # Remove the approved row from the suggestions tab
    ws.delete_rows(row_number, 1)
    wb.save(XLSX_FILE)
    return jsonify({"ok": True, "applied": new_kw})


@app.route("/api/keyword-suggestions/<int:row_number>", methods=["DELETE"])
@require_auth
def api_decline_kw_suggestion(row_number):
    """Remove a pending keyword suggestion without applying it."""
    if row_number < 2:
        return jsonify({"error": "Cannot delete header row"}), 400
    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "Workbook missing"}), 404
    wb = load_workbook(XLSX_FILE)
    if KW_SUGGEST_TAB not in wb.sheetnames:
        return jsonify({"error": "Suggestions tab missing"}), 404
    ws = wb[KW_SUGGEST_TAB]
    if row_number > ws.max_row:
        return jsonify({"error": "Row out of range"}), 404
    ws.delete_rows(row_number, 1)
    wb.save(XLSX_FILE)
    return jsonify({"ok": True})


@app.route("/api/source-requests/<int:row_number>/push", methods=["POST"])
@require_auth
def api_push_source_request(row_number):
    """Try to auto-apply a source request by writing to sources_config.json.
    Simple RSS + subreddit requests are auto-applicable. Complex API/social
    requests return 400 with a reason explaining that dev review is needed."""
    if row_number < 2:
        return jsonify({"error": "Cannot push header row"}), 400
    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "Workbook missing"}), 404

    wb = load_workbook(XLSX_FILE)
    if REQUESTS_TAB not in wb.sheetnames:
        return jsonify({"error": "Requests tab missing"}), 404
    ws = wb[REQUESTS_TAB]
    if row_number > ws.max_row:
        return jsonify({"error": "Row out of range"}), 404

    source_name = (ws.cell(row=row_number, column=1).value or "").strip()
    type_       = (ws.cell(row=row_number, column=2).value or "").strip().lower()
    url_target  = (ws.cell(row=row_number, column=3).value or "").strip()

    if not source_name:
        return jsonify({"error": "Row has no source name"}), 400

    cfg = _load_sources_config()
    applied_to = None

    if type_ == "rss":
        if not url_target or not url_target.lower().startswith(("http://", "https://")):
            return jsonify({"error": "RSS request has no valid URL. Edit the request or push manually."}), 400
        cfg["rss_feeds"].append({
            "name":     source_name,
            "url":      url_target,
            "category": "news",  # default category; admin can re-categorize later
        })
        applied_to = "RSS feeds"

    elif type_ == "forum":
        sub = _extract_subreddit(url_target) or _extract_subreddit(source_name)
        if not sub:
            return jsonify({
                "error": "Forum request doesn't look like a Reddit subreddit "
                         "(expected URL like reddit.com/r/X or 'r/X'). "
                         "For Hacker News or Stack Exchange additions, Claude needs to wire them up."
            }), 400
        existing = {s.lower() for s in cfg["reddit_subreddits"]}
        if sub.lower() in existing:
            return jsonify({"error": f"r/{sub} is already in the runtime subreddit list."}), 400
        cfg["reddit_subreddits"].append(sub)
        applied_to = f"Reddit subreddits (r/{sub})"

    else:
        return jsonify({
            "error": f"'{type_ or 'unknown'}' requests need dev review. Paste the request details to Claude."
        }), 400

    _save_sources_config(cfg)
    # Remove the pushed row from the queue
    ws.delete_rows(row_number, 1)
    wb.save(XLSX_FILE)
    return jsonify({"ok": True, "applied_to": applied_to})


@app.route("/api/source-requests/<int:row_number>", methods=["DELETE"])
@require_auth
def api_delete_source_request(row_number):
    if row_number < 2:
        return jsonify({"error": "Cannot delete header row"}), 400
    if not os.path.exists(XLSX_FILE):
        return jsonify({"error": "Workbook not found"}), 404

    wb = load_workbook(XLSX_FILE)
    if REQUESTS_TAB not in wb.sheetnames:
        return jsonify({"error": "Requests tab missing"}), 404
    ws = wb[REQUESTS_TAB]
    if row_number > ws.max_row:
        return jsonify({"error": "Row out of range"}), 404
    ws.delete_rows(row_number, 1)
    wb.save(XLSX_FILE)
    return jsonify({"ok": True})


# ============================================================================
# Sources (live stats) — read-only passthrough to sources.json
# ============================================================================

@app.route("/api/sources", methods=["GET"])
def api_sources():
    if not os.path.exists(SOURCES_JSON):
        return jsonify({"error": "No sources.json — run the scraper first."}), 404
    with open(SOURCES_JSON, "r", encoding="utf-8") as f:
        return f.read(), 200, {"Content-Type": "application/json"}


# ============================================================================
# Scheduled scraper
# ============================================================================
# Runs the scraper twice a day (at 08:00 and 18:00 UTC by default).
# Override with SCRAPE_HOURS env var (comma-separated UTC hours, e.g. "12,0").
# Set SCRAPE_ON_STARTUP=1 to run a scrape ~60s after boot so a fresh deploy
# has content without waiting for the next scheduled time.

def run_scrape_sync():
    """Invoke the scraper as a subprocess so it can't crash the web server."""
    print(f"[scrape] starting at {datetime.utcnow().isoformat()}Z")
    try:
        subprocess.run(
            [sys.executable, os.path.join(HERE, "scraper", "scraper.py")],
            check=False,
            timeout=60 * 20,  # 20 min hard cap
        )
        print(f"[scrape] finished at {datetime.utcnow().isoformat()}Z")
    except Exception as e:
        print(f"[scrape] ERROR: {e}")


def _start_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
    except ImportError:
        print("[scheduler] apscheduler not installed; skipping schedule.")
        return

    # Qubie lives in LA — run the cron on Pacific time so scrapes fire at the
    # same local hour year-round (no DST drift).
    scheduler = BackgroundScheduler(timezone="America/Los_Angeles", daemon=True)
    hours_str = os.environ.get("SCRAPE_HOURS", "7,16")
    for h in [h.strip() for h in hours_str.split(",") if h.strip()]:
        try:
            scheduler.add_job(run_scrape_sync, CronTrigger(hour=int(h), minute=0),
                              id=f"scrape-{h}", replace_existing=True)
            print(f"[scheduler] scrape scheduled at {h}:00 PT daily")
        except ValueError:
            print(f"[scheduler] invalid hour {h!r}, skipping")
    scheduler.start()

    # Optional warm-up scrape so a fresh deploy has content quickly
    if os.environ.get("SCRAPE_ON_STARTUP") == "1":
        threading.Timer(60, run_scrape_sync).start()
        print("[scheduler] warm-up scrape in 60s")


@app.route("/api/scrape", methods=["POST"])
@require_auth
def api_manual_scrape():
    """Manually trigger a scrape (useful from the admin pages)."""
    threading.Thread(target=run_scrape_sync, daemon=True).start()
    return jsonify({"ok": True, "status": "scrape started in background"})


# Kick off the scheduler when the module loads (works under gunicorn/flask run alike)
_start_scheduler()


# ============================================================================
# Main
# ============================================================================

if __name__ == "__main__":
    print(f"QUBIE News server starting on port {PORT}")
    print(f"  /           - Report")
    print(f"  /sources    - Sources admin")
    print(f"  /keywords   - Keywords admin")
    print(f"  DATA_DIR    = {DATA_DIR}")
    app.run(host="0.0.0.0", port=PORT, debug=False)
